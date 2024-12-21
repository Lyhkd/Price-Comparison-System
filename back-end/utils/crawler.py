import requests, json
from bs4 import BeautifulSoup
from time import sleep
import openpyxl
from urllib.parse import quote  
import aiohttp
import asyncio
import os
import time

class JDCrawler(object):
    def __init__(self, cookie, use_cookie=True):
        if use_cookie:
            self.headers = {
                'referer': 'https://search.jd.com/', 
                'accept' : 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'accept-encoding': 'gzip, deflate, br',
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36', 
                'cookie': cookie}
        else:
            self.headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/80.0.3987.149 Safari/537.36'
            }
        self.sess = requests.session()
        self.sess.headers.update(self.headers)
        
    def update_search(self, keyword, page_begin=1, page_end=1):
        start_url = 'https://search.jd.com/Search?keyword=' + keyword + '&enc=utf-8&wq=' + keyword
        self.url_list = [start_url + '&page=' + str(j) for j in range(page_begin, page_end + 1)]
    
            
    def get_item_info_dict(self):
        item_list = []  # 用于存储所有商品的字典
        for url in self.url_list:
            res = self.sess.get(url)
            res.encoding = 'utf-8'
            res = res.text
            # 定位搜索结果主体，并获取所有的商品的标签
            soup = BeautifulSoup(res, 'html.parser').select('#J_goodsList > ul')
            try:
                good_list = soup[0].select('[class=gl-item]')
            except:
                print('Skip No items found')
                continue
            # 循环获取所有商品信息
            for temp in good_list:
                item_info = {}  # 用于存储单个商品的字典
                
                # sku
                sku = temp.get('data-sku').strip()
                item_info['sku'] = sku
                
                # 获取名称信息
                name_div = temp.select_one('[class="p-name p-name-type-2"]').find('em')
                name = name_div.text.replace('\n', '').replace('\t', '').strip()
                
                item_info['title'] = name
                
                # 商品链接
                item_url = temp.select_one('[class=p-img]').select_one('a').get('href')
                item_info['link'] = "https:" + item_url

                # 价格信息
                price_div = temp.select_one('[class=p-price]')
                item_info['price'] = price_div.text.strip()[1:]

                # 店铺信息
                shop_div = temp.select_one('[class=p-shop]')
                item_info['shop'] = shop_div.get_text().strip()
                item_info['shop_link'] = "https:" + shop_div.select_one('a').get('href')
                
                # 图片信息
                img_div = temp.select_one('[class=p-img]').select_one('img')
                img_url = "https:" + img_div.get('data-lazy-img')
                item_info['img_url'] = img_url
            

                # 将商品信息字典加入列表
                item_list.append(item_info)
            sleep(1)
        # 返回商品字典列表
        return item_list
    
    def get_detail(self, sku=None):
        commit_start_url = f'https://item.jd.com/{sku}.html'
        # 发送请求，得到结果
        res = self.sess.get(commit_start_url)
        bs = BeautifulSoup(res.text, 'html.parser')
        # save_html(res.text)
        # price = bs.select('[classp=p-price]')
        # print("try to get detail")
        # print(price)
        paras = bs.select('[class=p-parameter]')
        pics = bs.select('[class=spec-items]')
        if len(pics) != 0:
            pictures = pics[0].select('img')
        else:
            pictures = []
        parameters = paras[0].select('li')
        # brand = soup[0].select('li')[0].get('title')
        attr = {}
        for parameter in parameters:
            attr[parameter.get_text().split('：')[0]] = parameter.get_text().split('：')[1].strip()
        attr['img_list'] = ""
        for pic in pictures:
            attr['img_list'] += pic.get('src')+";"
        return attr
    
    def get_detail_string(self, sku):
        attrs = self.get_detail(sku) 
        if len(attrs.keys()) == 0:
            return None
        # print(attrs)
        results = '' # attrs是一个dict，stringfy之后返回
        for (key, value) in attrs.items():
            if value is not None and key != '店铺':
                results += key+':'+value+'\n'
        return results

    def read_good_info_xlsx(self, file_path):
        """
        从xlsx文件中读取商品信息，并返回字典列表。

        :param file_path: xlsx文件路径
        :return: 包含商品信息的字典列表
        """
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active  # 默认读取活动工作表
        
        # 检查表头
        headers = [cell.value for cell in ws[1]]  # 获取第一行作为表头
        expected_headers = ["sku", "title", "link", "price", "shop", "shop_link", "img_url"]
        if headers != expected_headers:
            raise ValueError(f"表头与预期不符，请检查文件格式！预期表头为: {expected_headers}")

        # 初始化商品信息字典列表
        item_list = []

        # 从第二行开始读取数据
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 将每一行数据转换为字典
            item_info = {
                "sku": row[0],
                "title": row[1].replace('\n', '').replace('\t', '').strip(),
                "link": row[2],
                "price": row[3],
                "shop": row[4],
                "shop_link": row[5],
                "img_url": row[6],
            }
            item_list.append(item_info)
        
        return item_list


class GWCrawler(object):
    def __init__(self, cookie, use_cookie=True):
        if use_cookie:
            self.headers = {
                'authority': 'www.gwdang.com',
                'referer': 'https://www.gwdang.com/',
                'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
                'accept-encoding': 'gzip, deflate, br, zstd',
                'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36',
                'sec-fetch-site': 'same-origin',
                'sec-fetch-mode': 'navigate',
                'sec-fetch-user': '?1',
                'sec-fetch-dest': 'document',
                'accept-language': 'zh-CN,zh;q=0.9,en;q=0.8',
                'cookie': cookie
            }
        else:
            self.headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                              'Chrome/80.0.3987.149 Safari/537.36'
            }
        self.PLATFORM = {
            'JD': '3',
            'TM': '83',
            'SN': '25'
        }
        
    async def create_session(self):
        self.session = aiohttp.ClientSession(headers=self.headers)
        
    async def close_session(self):
        if self.session:
            await self.session.close()
            
    def update_search(self, keyword, platform='JD', page_begin=1, page_end=1):
        print(keyword, platform, page_begin, page_end)
        platform_str = ''
        if platform in self.PLATFORM:
            platform_str = self.PLATFORM[platform]
        start_url = 'https://www.gwdang.com/search?crc64='
        encoded_keyword = quote(keyword) 
        self.url_list = [start_url + str(j) + '&s_product=' + encoded_keyword + '&site_id=' + platform_str for j in range(page_begin, page_end + 1)]
        self.end_page = page_end
        
    async def fetch(self, url):
        async with self.session.get(url) as response:
            response_text = await response.text()
            # save_html(response_text)
            return response_text

    async def fetch_all(self):
        tasks = [self.fetch(url) for url in self.url_list]
        return await asyncio.gather(*tasks)

    
    # 获取商品详情信息
    def get_item_detail_info(self, html):
        soup = BeautifulSoup(html, 'html.parser')
        item_info = {}
        # 获取商品图片
        info_div = soup.select_one('[class=product-info-table]')
        if info_div:
            infos = info_div.select('[class=info-item]')
            for info in infos:
                key = info.select_one(".info-key").text.replace(":","").strip()
                value = info.select_one(".info-value").text.strip()
                item_info[key] = value
        # 获取商品描述
        return item_info

    async def get_detail_string(self, sku=None, platform='JD'):
        if platform in self.PLATFORM:
            platform_str = self.PLATFORM[platform]
        else:
            return None
        item_url = f'https://www.gwdang.com/crc64/dp{sku}-{platform_str}'
        print("in get detail string", item_url)
        await self.create_session()
        async with self.session.get(item_url) as response:
            raw_html = await response.read()  # 直接读取响应的字节内容
            try:
                html = raw_html.decode('utf-8')  # 尝试使用 UTF-8 解码
            except UnicodeDecodeError:
                html = raw_html.decode('gbk')  # 如果 UTF-8 解码失败，尝试使用 GBK 解码
            # save_html(html)
            item_info = self.get_item_detail_info(html)
            item_info_str = ''
            for key, value in item_info.items():
                item_info_str += key + ':' + value + '\n'
            await self.close_session()
            return item_info_str
    
    async def get_item_info_dict(self):
        item_list = []  # 用于存储所有商品的字典
        await self.create_session()
        results = await self.fetch_all()
        await self.close_session()
        sleep(5)
        for res in results:
            soup = BeautifulSoup(res, 'html.parser').select('[class="dp-list"]')
            if not soup:
                continue
            good_list = soup[0].select('li')
            print("get good_list", len(good_list))
            for temp in good_list:
                if temp.get('data-dp-id') is None:
                    continue
                item_info = {}  # 用于存储单个商品的字典

                # sku
                sku = temp.get('data-dp-id').strip().split('-')[0]
                item_info['sku'] = sku

                # 获取名称信息
                name_div = temp.select_one('[class="item-title"]')
                name = name_div.text.strip()
                item_info['title'] = name

                # 商品链接
                commit_start_url = f'https://item.jd.com/{sku}.html'
                item_info['link'] = commit_start_url

                # 价格信息
                price_div = temp.select_one('[class=bigRedPrice]')
                item_info['price'] = price_div.text.strip()[1:]

                # 店铺信息
                shop_div = temp.select_one('[class=site]')
                item_info['shop'] = shop_div.get_text().strip()

                # 图片信息
                img_url = temp.select_one('[class=item-img]').select_one('img').get('data-original')
                item_info['img_url'] = img_url

                # 将商品信息字典加入列表
                item_list.append(item_info)

        # 返回商品字典列表
        return item_list
    
class AmazonCrawler(object):
    def __init__(self, cookie):
        self.cookie = cookie
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/128.0.0.0 Safari/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Connection": "keep-alive",
            "Host": "www.amazon.com",
            "Pragma": "no-cache",
            "Cache-Control": "no-cache",
            "device-memory": "8",
            "sec-ch-device-memory": "8",
            "dpr": "1",
            "sec-ch-dpr": "1",
            "viewport-width": "1287",
            "sec-ch-viewport-width": "1287",
            "rtt": "250",
            "downlink": "10",
            "ect": "4g",
            "sec-ch-ua": '"Chromium";v="128", "Not;A=Brand";v="24", "Google Chrome";v="128"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-ch-ua-platform-version": '"15.0.0"',
            "Upgrade-Insecure-Requests": "1",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-User": "?1",
            "Sec-Fetch-Dest": "document",
            "Accept-Language": "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7,fr;q=0.6",
            "Cookie": self.cookie,   # 直接从你设置邮编后的亚马逊网页 F12 查看获取
        }
        
    def update_search(self, keyword, page_begin=1, page_end=1):
        # 监控的产品，搜索词
        search_url = f"https://www.amazon.com/s?k={keyword.replace(' ', '+')}"
        self.url_list = [search_url + '&page=' + str(j) for j in range(page_begin, page_end + 1)]
    
    async def fetch_page(self, session, url):
        async with session.get(url, headers=self.headers) as response:
            html = await response.text()
            await asyncio.sleep(2)  # 延迟2秒
            if "dogs of amazon" in html.lower():
                print("搜索被标识为异常访问")
            return self.get_item_info_dict(html)
            
    async def get_item_amazon(self):
        async with aiohttp.ClientSession() as session:
            tasks = [self.fetch_page(session, url) for url in self.url_list]
            results = await asyncio.gather(*tasks)
            # save_html(html)
            
            item_list = [item for sublist in results for item in sublist]
            # print(item_list)
            await session.close()
            return item_list
        
    def get_item_info_dict(self, html):
        item_list = []  # 用于存储所有商品的字典
        # 定位搜索结果主体，并获取所有的商品的标签
        soup = BeautifulSoup(html, 'html.parser')
        try:
            good_list = soup.select('div.s-main-slot div.s-result-item')
        except:
            print('Skip No items found')
        # 循环获取所有商品信息
        for temp in good_list:
            item_info = {}  # 用于存储单个商品的字典
            
            # sku
            asin = temp.get('data-asin')
            if not asin:
                continue
            item_info['sku'] = asin
            
            # 获取名称信息
            name_div = temp.select_one('div[data-cy="title-recipe"] a h2 span')
            if name_div:
                item_info['title'] = name_div.text.replace('\n','').strip()
                if len(item_info['title']) >= 200:
                    item_info['title'] = item_info['title'][:200]
            
            # 商品链接
            item_url = temp.select_one('a.a-link-normal')
            if item_url:
                item_info['link'] = f"https://www.amazon.com/zh/dp/{asin}"
            else:
                continue
            # 价格信息
            price_whole = temp.select_one('span.a-price-whole')
            price_fraction = temp.select_one('span.a-price-fraction')
            if price_whole and price_fraction:
                item_info['price'] = price_whole.text.replace(',','').strip() + price_fraction.text.strip()
            else:
                item_info['price'] = None
            
            # 图片信息
            img_div = temp.select_one('img.s-image')
            if img_div:
                item_info['img_url'] = img_div.get('src')
            
            item_info['shop'] = "Amazon"
            item_info['shop_link'] = "https://www.amazon.com"
            # 将商品信息字典加入列表
            item_list.append(item_info)
        # 返回商品字典列表
        return item_list
    
    
    def get_item_detail_info(self, html):
        soup = BeautifulSoup(html, 'html.parser')
        item_info = {}
        save_html(html)

        # 获取商品图片
        img_div = soup.select_one('#imgTagWrapperId img')
        if img_div:
            item_info['img_list'] = img_div.get('src').replace('https:', '')  # 去掉 https: 部分

        # 获取商品描述
        description_div = soup.select_one('#productDescription')
        if description_div:
            item_info['简介'] = description_div.text.strip()
        detail = []
        detail_infos = []
        detail_infos.append(soup.select_one('#productFactsDesktopExpander'))
        detail_infos.append(soup.select_one('#detailBullets_feature_div'))
        detail_infos.append(soup.select_one('#detailBulletsWrapper_feature_div'))
        for detail_info in detail_infos:
            if detail_info:
                detail_info = detail_info.select('li')
                if detail_info is not None:
                    detail += detail_info
        # print(detail)
        for i in detail:
            if ':' in i.text:
                key, value = i.text.split(':', 1)  # 只分割一次，避免值中有 ":" 的情况
                item_info[key.replace('\n', '').replace('\u200f', '').strip()] = value.replace('\n', '').replace('\u200e', '').strip()
            else:
                item_info['描述'] = i.text
                break
                
            
        
        # 获取商品评分
        rating_div = soup.select_one('.reviewCountTextLinkedHistogram')
        if rating_div:
            item_info['评分'] = rating_div.select_one('[class="a-size-base a-color-base"]').text.strip()
    

        # 获取商品评论数
        review_count_div = soup.select_one('#acrCustomerReviewText')
        if review_count_div:
            item_info['评论数'] = review_count_div.text.strip()

        # 转换成字符串
        item_info_str = ''
        for key, value in item_info.items():
            item_info_str += key + ':' + value + '\n'
        
        return item_info_str

    async def fetch_item_detail(self,session, url):
        async with session.get(url, headers=self.headers) as response:
            html = await response.text()
            return self.get_item_detail_info(html)
        
    async def get_detail_string(self, sku=None):
        item_url = f"https://www.amazon.com/zh/dp/{sku}"  # 替换为实际的商品详情页 URL
        print("in get detail string", item_url)
        async with aiohttp.ClientSession() as session:
            item_info_str = await self.fetch_item_detail(session, item_url)
            await session.close()
            return item_info_str
        
    async def get_item_price(self, sku=None):
        item_url = f"https://www.amazon.com/zh/dp/{sku}"
        async  with aiohttp.ClientSession() as session:
            async with session.get(item_url, headers=self.headers) as response:
                html = await response.text()
                await session.close()
                return self.get_price(html)

def save_html(content):
    # 确保目录存在
    directory = "page"
    if not os.path.exists(directory):
        os.makedirs(directory)
 
    # 生成文件名
    filename = time.strftime("%Y%m%d%H%M%S") + ".html"
    filepath = os.path.join(directory, filename)
 
    # 保存文件
    with open(filepath, 'w', encoding='utf-8') as file:
        file.write(content)
    print("页面已保存至:", filepath)
    
    
if __name__ == "__main__":
    # cookie，用于验证登录状态，必须要有cookie，否则京东会提示网络繁忙请重试
    # 获取方法：使用浏览器登录过后按F12，点击弹出界面中最上方的network选项，name栏里面随便点开一个，拉到最下面就有cookie，复制到cookie.txt中
    # 注意，不要换行，前后不要有空格，只需要复制cookie的值，不需要复制“cookie：”这几个字符
    # 上面的看不懂的话，看这个：https://blog.csdn.net/qq_46047971/article/details/121694916
    # 然后就可以运行程序了
    cookie_str = ''
    with open('back-end/data/cookie.txt') as f:
        cookie_str = f.readline()
    
    # 输入cookie，关键词，输入结束页数
    content_page = JDCrawler(cookie_str)
    content_page.update_search('手机', 1)
    
    content_page.print()
    print(content_page.to_json())
    # content_page.get_item_info_xslx()
